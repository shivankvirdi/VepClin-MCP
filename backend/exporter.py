import csv
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from unidecode import unidecode

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

_BOLD_RE = re.compile(r"\*\*(.+?)\*\*")
_TYPOGRAPHIC_REPLACEMENTS = {
    "\u2014": "-",   # em dash —
    "\u2013": "-",   # en dash –
    "\u2192": "->",  # arrow →
    "\u2018": "'", "\u2019": "'",   # curly single quotes
    "\u201c": '"', "\u201d": '"',   # curly double quotes
    "\u00b2": "2",   # superscript 2 (as in "PolyPhen²")
    "\u2022": "-",   # bullet character, when used inline rather than as a real list marker
}

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="3C3836", end_color="3C3836", fill_type="solid")

IMPACT_FILLS = {
    "HIGH": PatternFill(start_color="FB4934", end_color="FB4934", fill_type="solid"),
    "MODERATE": PatternFill(start_color="FABD2F", end_color="FABD2F", fill_type="solid"),
}

_BASE_STYLES = getSampleStyleSheet()
HEADER_CELL_STYLE = ParagraphStyle(
    "TableHeaderCell",
    parent=_BASE_STYLES["BodyText"],
    textColor=colors.white,
    fontName="Helvetica-Bold",
)

def _sanitize_pdf_text(text: str) -> str:
    return unidecode(text)


def _convert_inline_markdown(text: str) -> str:
    text = _sanitize_pdf_text(text)
    return _BOLD_RE.sub(r"<b>\1</b>", text)


def _markdown_table_to_flowable(lines: list[str], cell_style) -> Table:
    rows = []
    for line in lines:
        if re.match(r"^\|?\s*:?-+:?\s*(\|\s*:?-+:?\s*)+\|?$", line):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        style_to_use = HEADER_CELL_STYLE if not rows else cell_style
        rows.append([Paragraph(_convert_inline_markdown(c), style_to_use) for c in cells])
    return _styled_table(rows)

def _markdown_to_flowables(text: str, styles) -> list:
    """Convert a chunk of LLM-generated Markdown into a list of reportlab
    flowables: headings, bold-aware paragraphs, bullet points, and tables.

    This only covers the subset of Markdown a chat model typically produces
    (headers, bold, bullets, simple pipe tables) — it is not a general
    Markdown parser.
    """
    flowables = []
    lines = text.strip().split("\n")
    i = 0

    while i < len(lines):
        line = lines[i].strip()

        if not line:
            i += 1
            continue

        if line.startswith("### "):
            flowables.append(Paragraph(_convert_inline_markdown(line[4:]), styles["Heading4"]))
            i += 1
            continue
        if line.startswith("## "):
            flowables.append(Paragraph(_convert_inline_markdown(line[3:]), styles["Heading3"]))
            i += 1
            continue
        if line.startswith("# "):
            flowables.append(Paragraph(_convert_inline_markdown(line[2:]), styles["Heading2"]))
            i += 1
            continue

        if line.startswith("|"):
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                table_lines.append(lines[i])
                i += 1
            flowables.append(_markdown_table_to_flowable(table_lines, styles["BodyText"]))
            flowables.append(Spacer(1, 0.1 * inch))
            continue

        if line.startswith(("- ", "• ", "* ")):
            bullet_text = _convert_inline_markdown(line[2:].strip())
            flowables.append(Paragraph(f"• {bullet_text}", styles["BodyText"]))
            i += 1
            continue

        flowables.append(Paragraph(_convert_inline_markdown(line), styles["BodyText"]))
        i += 1

    return flowables

def _flatten_result(row: dict) -> dict:
    """Flatten one batch result (VEP fields + nested ClinVar match) into a
    single flat dict suitable for one delimited-file row.
    """
    clinvar = row.get("clinvar", {})
    matches = clinvar.get("matches") or []
    top = matches[0] if matches else {}

    return {
        "input": row.get("input"),
        "gene_symbol": row.get("gene_symbol"),
        "consequence": row.get("consequence"),
        "protein_change_short": row.get("protein_change_short"),
        "impact": row.get("impact"),
        "sift_prediction": row.get("sift_prediction"),
        "polyphen_prediction": row.get("polyphen_prediction"),
        "clinvar_found": clinvar.get("found", False),
        "germline_classification": top.get("germline_classification"),
        "germline_review_status": top.get("germline_review_status"),
        "oncogenicity_classification": top.get("oncogenicity_classification"),
        "clinical_impact_classification": top.get("clinical_impact_classification"),
        "clinvar_id": top.get("variation_id"),
    }

def export_variant_report_pdf(
    vep_result: dict | None,
    clinvar_result: dict | None,
    summary_text: str | None,
    path: str,
) -> None:
    """Generate a single-variant PDF report from the most recent chat lookup.

    Either argument may be None — a user can look up ClinVar directly for a
    well-known variant (e.g. 'BRAF V600E') without ever calling VEP.
    """
    if not vep_result and not clinvar_result:
        raise ValueError("No lookup data to report on. Look up a variant in chat first.")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], textColor=colors.HexColor("#282828")
    )
    heading_style = ParagraphStyle(
        "SectionHeading", parent=styles["Heading2"], textColor=colors.HexColor("#8ec07c")
    )

    story = []

    gene = (vep_result or {}).get("gene_symbol") or "Unknown gene"
    protein = (vep_result or {}).get("protein_change_short") or ""
    story.append(Paragraph(f"Variant Report: {gene} {protein}".strip(), title_style))
    story.append(Spacer(1, 0.25 * inch))

    # ---- VEP section ----
    story.append(Paragraph("Ensembl VEP Consequence", heading_style))
    if vep_result is None:
        story.append(Paragraph("No VEP lookup was performed for this variant.", styles["BodyText"]))
    elif vep_result.get("error"):
        story.append(Paragraph(f"Lookup error: {vep_result.get('message', 'Unknown error')}", styles["BodyText"]))
    else:
        body_style = styles["BodyText"]
        vep_rows = [
            _table_row(["Field", "Value"], HEADER_CELL_STYLE),
            _table_row(["Gene", vep_result.get("gene_symbol") or "—"], body_style),
            _table_row(["Consequence", (vep_result.get("consequence") or "—").replace("_", " ")], body_style),
            _table_row(["Protein change", vep_result.get("protein_change") or "—"], body_style),
            _table_row(["Impact", vep_result.get("impact") or "—"], body_style),
            _table_row(["SIFT", vep_result.get("sift_prediction") or "—"], body_style),
            _table_row(["PolyPhen", vep_result.get("polyphen_prediction") or "—"], body_style),
            _table_row(["Genome build", vep_result.get("build") or "—"], body_style),
        ]
        story.append(_styled_table(vep_rows))

    story.append(Spacer(1, 0.25 * inch))

   # ---- ClinVar section ----
    story.append(Paragraph("ClinVar Findings", heading_style))
    story.append(Spacer(1, 0.1 * inch))
    if clinvar_result is None:
        story.append(Paragraph("No ClinVar lookup was performed for this variant.", styles["BodyText"]))
    elif not clinvar_result.get("found"):
        story.append(Paragraph("No ClinVar matches were found for this variant.", styles["BodyText"]))
    else:
        for i, match in enumerate(clinvar_result.get("matches", []), start=1):
            story.append(Paragraph(f"Match {i}: {match.get('title') or 'Untitled record'}", styles["Heading3"]))
            story.append(Paragraph(f"ClinVar Variation ID: {match.get('variation_id') or '—'}", styles["BodyText"]))
            story.append(Spacer(1, 0.1 * inch))
            
            classification_blocks = [
                ("Germline Classification", "germline_classification", "germline_review_status",
                 "germline_last_evaluated", "germline_traits"),
                ("Clinical Impact (Somatic/Oncology)", "clinical_impact_classification",
                 "clinical_impact_review_status", "clinical_impact_last_evaluated", "clinical_impact_traits"),
                ("Oncogenicity", "oncogenicity_classification", "oncogenicity_review_status",
                 "oncogenicity_last_evaluated", "oncogenicity_traits"),
            ]

            for label, class_key, review_key, date_key, traits_key in classification_blocks:
                classification = match.get(class_key)
                if not classification:
                    continue  # skip sections with no data, same as the terminal display does

                traits = match.get(traits_key) or []
                story.append(Paragraph(label, styles["Heading4"]))
                block_rows = [
                    _table_row(["Field", "Value"], HEADER_CELL_STYLE),
                    _table_row(["Classification", classification], body_style),
                    _table_row(["Review status", match.get(review_key) or "—"], body_style),
                    _table_row(["Last evaluated", match.get(date_key) or "—"], body_style),
                    _table_row(["Traits", ", ".join(str(t) for t in traits) if traits else "—"], body_style),
                ]
                story.append(_styled_table(block_rows))
                story.append(Spacer(1, 0.1 * inch))

            if match.get("position_verified") is not None:
                verified_text = "Yes" if match["position_verified"] else "No — possible ambiguity, review carefully"
                story.append(Paragraph(f"Genomic position verified against VEP: {verified_text}", styles["BodyText"]))

            story.append(Spacer(1, 0.2 * inch))

    # ---- LLM-generated summary, if available ----
    if summary_text:
        story.append(Paragraph("Summary", heading_style))
        story.extend(_markdown_to_flowables(summary_text, styles))
        story.append(Spacer(1, 0.2 * inch))
    
    doc = SimpleDocTemplate(path, pagesize=letter)
    doc.build(story)

def _table_row(cells: list[str], style) -> list:
    return [Paragraph(_sanitize_pdf_text(str(c)), style) for c in cells]

PAGE_CONTENT_WIDTH = letter[0] - 2 * inch

def _styled_table(rows: list, col_widths: list | None = None) -> Table:
    num_cols = len(rows[0]) if rows else 1
    if col_widths is None:
        col_widths = [PAGE_CONTENT_WIDTH / num_cols] * num_cols

    table = Table(rows, colWidths=col_widths, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3c3836")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#928374")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f4f4")]),
    ]))
    return table

def _style_header_row(ws, num_columns: int) -> None:
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(num_columns)}1"


def _autosize_columns(ws, num_columns: int) -> None:
    for col in range(1, num_columns + 1):
        letter = get_column_letter(col)
        max_len = max(
            (len(str(cell.value)) for cell in ws[letter] if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 50)


def export_to_xlsx(results: list[dict], path: str) -> None:
    wb = Workbook()

    # ---- Sheet 1: one row per variant, matching the CSV/TSV summary shape ----
    summary_ws = wb.active
    summary_ws.title = "Variant Summary"

    summary_headers = list(_flatten_result(results[0]).keys()) if results else []
    summary_ws.append(summary_headers)

    for row in results:
        flat = _flatten_result(row)
        summary_ws.append([flat.get(h) for h in summary_headers])

        impact = flat.get("impact")
        if impact in IMPACT_FILLS:
            for col in range(1, len(summary_headers) + 1):
                summary_ws.cell(row=summary_ws.max_row, column=col).fill = IMPACT_FILLS[impact]

    _style_header_row(summary_ws, len(summary_headers))
    _autosize_columns(summary_ws, len(summary_headers))

    # ---- Sheet 2: every ClinVar match for every variant, nothing dropped ----
    detail_ws = wb.create_sheet("ClinVar Detail")
    detail_headers = [
        "input", "variation_id", "title",
        "germline_classification", "germline_review_status",
        "clinical_impact_classification", "oncogenicity_classification",
        "traits", "position_verified",
    ]
    detail_ws.append(detail_headers)

    for row in results:
        clinvar = row.get("clinvar", {})
        for match in clinvar.get("matches", []):
            traits = (
                match.get("clinical_impact_traits")
                or match.get("oncogenicity_traits")
                or match.get("germline_traits")
                or []
            )
            detail_ws.append([
                row.get("input"),
                match.get("variation_id"),
                match.get("title"),
                match.get("germline_classification"),
                match.get("germline_review_status"),
                match.get("clinical_impact_classification"),
                match.get("oncogenicity_classification"),
                ", ".join(str(t) for t in traits),
                match.get("position_verified"),
            ])

    _style_header_row(detail_ws, len(detail_headers))
    _autosize_columns(detail_ws, len(detail_headers))

    if not results:
        raise ValueError("No results to export.")

    wb.save(path)

def _export_delimited(results: list[dict], path: str, delimiter: str) -> None:
    rows = [_flatten_result(r) for r in results]

    if not rows:
        raise ValueError("No results to export.")

    fieldnames = list(rows[0].keys())

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delimiter)
        writer.writeheader()
        writer.writerows(rows)


def export_to_csv(results: list[dict], path: str) -> None:
    _export_delimited(results, path, delimiter=",")


def export_to_tsv(results: list[dict], path: str) -> None:
    _export_delimited(results, path, delimiter="\t")

def export_annotated_vcf(original_vcf_path: str, results: list[dict], output_path: str) -> dict:
    """Write a new VCF with VEP/ClinVar annotation added to the INFO column
    as a CSQ= field, following the same convention the standalone VEP tool uses.

    Multi-allelic rows (comma-separated ALT) are annotated per-allele, each
    allele's CSQ block separated by a comma, matching VEP's own convention
    for representing multiple outcomes at one position.

    Returns a dict with counts: annotated, no_match (rows with too few
    columns to identify a variant, or a variant we have no result for),
    and passthrough (header/blank lines, unaffected).
    """
    lookup: dict[tuple[str, str, str, str], dict] = {}
    for result in results:
        input_str = result.get("input", "")
        parts = input_str.split()
        if len(parts) < 5:
            continue
        chrom, pos, _id, ref, alt = parts[0], parts[1], parts[2], parts[3], parts[4]
        lookup[(chrom, pos, ref, alt)] = _flatten_result(result)

    csq_header = (
        '##INFO=<ID=CSQ,Number=.,Type=String,Description='
        '"Consequence annotations from VepClin-MCP. Format: '
        'Gene|Consequence|Impact|ProteinChange|SIFT|PolyPhen|'
        'ClinVar_Germline|ClinVar_ReviewStatus">\n'
    )

    stats = {"annotated": 0, "no_match": 0, "skipped": 0}

    with open(original_vcf_path, "r") as infile, open(output_path, "w") as outfile:
        header_written = False

        for line in infile:
            if line.startswith("##"):
                outfile.write(line)
                continue

            if line.startswith("#CHROM"):
                if not header_written:
                    outfile.write(csq_header)
                    header_written = True

                header_fields = line.rstrip("\n").split("\t")
                standard_cols = ["CHROM", "POS", "ID", "REF", "ALT", "QUAL", "FILTER", "INFO"]
                while len(header_fields) < len(standard_cols):
                    header_fields.append(standard_cols[len(header_fields)])
                outfile.write("\t".join(header_fields) + "\n")
                continue

            stripped = line.strip()
            if not stripped:
                outfile.write(line)
                continue

            fields = stripped.split("\t")
            if len(fields) < 5:
                # Not enough columns to even identify chrom/pos/ref/alt.
                stats["skipped"] += 1
                outfile.write(line)
                continue

            while len(fields) < 8:
                fields.append(".")

            chrom = fields[0].removeprefix("chr")
            pos, ref, alt = fields[1], fields[3], fields[4]

            csq_blocks = []
            row_had_match = False
            for single_alt in alt.split(","):
                flat = lookup.get((chrom, pos, ref, single_alt))
                if flat is None:
                    csq_blocks.append("")
                    continue
                row_had_match = True
                csq_blocks.append(
                    "|".join(
                        str(flat.get(key) or "") for key in (
                            "gene_symbol", "consequence", "impact",
                            "protein_change_short", "sift_prediction",
                            "polyphen_prediction", "germline_classification",
                            "germline_review_status",
                        )
                    )
                )

            if row_had_match:
                stats["annotated"] += 1
            else:
                stats["no_match"] += 1

            csq_value = ",".join(csq_blocks)
            info = fields[7]
            fields[7] = f"{info};CSQ={csq_value}" if info != "." else f"CSQ={csq_value}"

            outfile.write("\t".join(fields) + "\n")

    return stats